"""삼성카드 홈페이지에서 내려받은 이용내역 엑셀(.xlsx)을 DB로 임포트.

사용:
  pip install openpyxl
  python excel_import.py 이용내역.xlsx

헤더 행을 자동 탐지한다: '이용일'/'승인일', '가맹점', '금액' 계열 컬럼명이
모두 포함된 첫 행을 헤더로 본다. 실제 파일 컬럼명이 다르면 KEYWORDS만 고치면 된다.
"""

from __future__ import annotations

import re
import sys
from datetime import datetime
from pathlib import Path

import db
from parser import Transaction

KEYWORDS = {
    "date": ("이용일", "승인일", "거래일"),
    "merchant": ("가맹점", "이용가맹점", "가맹점명"),
    "amount": ("이용금액", "승인금액", "금액"),
    "installment": ("할부", "이용구분", "결제방법"),
    "status": ("상태", "승인구분", "취소"),
}


def _find_columns(rows: list[tuple]) -> tuple[int, dict[str, int]]:
    """(헤더 행 인덱스, {필드: 컬럼 인덱스})를 반환."""
    for i, row in enumerate(rows[:20]):
        cells = [str(c).strip() if c is not None else "" for c in row]
        cols: dict[str, int] = {}
        for field, keys in KEYWORDS.items():
            for j, cell in enumerate(cells):
                if any(k in cell for k in keys):
                    cols[field] = j
                    break
        if {"date", "merchant", "amount"} <= cols.keys():
            return i, cols
    raise ValueError("헤더 행을 찾지 못했습니다. excel_import.py의 KEYWORDS를 파일 컬럼명에 맞게 수정하세요.")


def _parse_date(value) -> datetime | None:
    if isinstance(value, datetime):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d", "%Y.%m.%d", "%Y/%m/%d", "%Y%m%d"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def _parse_amount(value) -> int | None:
    if isinstance(value, (int, float)):
        return int(value)
    digits = re.sub(r"[^\d-]", "", str(value))
    return int(digits) if digits and digits != "-" else None


def load_transactions(path: Path) -> list[Transaction]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        sys.exit("openpyxl이 필요합니다: pip install openpyxl")

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header_idx, cols = _find_columns(rows)

    txs: list[Transaction] = []
    for row in rows[header_idx + 1:]:
        ts = _parse_date(row[cols["date"]]) if row[cols["date"]] is not None else None
        amount = _parse_amount(row[cols["amount"]]) if row[cols["amount"]] is not None else None
        merchant = str(row[cols["merchant"]] or "").strip()
        if ts is None or amount is None or not merchant:
            continue

        installment = "일시불"
        if "installment" in cols and row[cols["installment"]]:
            installment = str(row[cols["installment"]]).strip() or "일시불"
        canceled = False
        if "status" in cols and row[cols["status"]]:
            canceled = "취소" in str(row[cols["status"]])
        # 취소 건이 음수 금액으로 오는 파일 대응
        if amount < 0:
            amount, canceled = -amount, True

        txs.append(Transaction(ts=ts, merchant=merchant, amount=amount,
                               installment=installment, canceled=canceled,
                               raw=f"excel:{path.name}"))
    return txs


def main() -> None:
    if len(sys.argv) != 2:
        sys.exit("사용법: python excel_import.py <이용내역.xlsx>")
    path = Path(sys.argv[1])
    txs = load_transactions(path)
    conn = db.connect()
    try:
        added, skipped = db.insert_many(conn, txs, source="excel")
    finally:
        conn.close()
    print(f"{path.name}: {len(txs)}건 읽음 → 신규 {added}건 저장, 중복 {skipped}건 스킵")


if __name__ == "__main__":
    main()
